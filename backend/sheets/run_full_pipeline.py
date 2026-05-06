"""
Single entrypoint to run PriyaFil ingestion + calculation pipeline.

Supports month-folder based inputs:
  InputTallyTarka/<month>/

Expected standard files inside month folder:
  - Sales_Pur_Exps*.xlsx  (contains Purchase, Inventory Sales, Direct/Indirect Expns, Stock Valuation)
  - Balance file can be either:
      1) MIS-style file with sheet "Balance Sheet"
      2) BS_PL_CF-style file with sheet like "BS_Apr25-Jan26" (single as-at month)
  - Salary file: PFCO*Salary*.xls or PFCO*Salary*.xlsx (OLE2 content supported by xlrd)
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Iterable

import openpyxl

if __package__ is None or __package__ == "":
    # Support direct execution: python sheets/run_full_pipeline.py
    sys.path.append(str(Path(__file__).resolve().parents[1]))

from sheets.db import MONTHS, get_connection


def _normalize_month_token(token: str) -> str:
    token = token.strip()
    if not token:
        raise ValueError("Empty month token")
    month_map = {
        "apr": "Apr",
        "april": "Apr",
        "may": "May",
        "jun": "Jun",
        "june": "Jun",
        "jul": "Jul",
        "july": "Jul",
        "aug": "Aug",
        "august": "Aug",
        "sep": "Sep",
        "sept": "Sep",
        "september": "Sep",
        "oct": "Oct",
        "october": "Oct",
        "nov": "Nov",
        "november": "Nov",
        "dec": "Dec",
        "december": "Dec",
        "jan": "Jan",
        "january": "Jan",
        "feb": "Feb",
        "february": "Feb",
        "mar": "Mar",
        "march": "Mar",
    }
    key = token.lower()
    if key not in month_map:
        raise ValueError(f"Invalid month: {token}")
    return month_map[key]


def _parse_months_arg(months_arg: str | None) -> list[str] | None:
    if not months_arg:
        return None
    months = [_normalize_month_token(x) for x in months_arg.split(",") if x.strip()]
    deduped = []
    for m in MONTHS:
        if m in months and m not in deduped:
            deduped.append(m)
    if not deduped:
        raise ValueError("No valid months parsed from --months")
    return deduped


def _find_first(folder: Path, patterns: Iterable[str]) -> Path | None:
    for pattern in patterns:
        matches = sorted(folder.glob(pattern))
        if matches:
            return matches[0]
    return None


def _ingest_balance_from_bs_pl_cf(
    conn, balance_file: Path, fy_suffix: str, month: str
) -> int:
    """Ingest a single-month balance sheet from BS_PL_CF-style workbook."""
    from sheets.ingest_balance_sheet import create_balance_sheet_table

    table_name = f"balance_sheet_{fy_suffix}"
    create_balance_sheet_table(conn, fy_suffix)

    wb = openpyxl.load_workbook(balance_file, read_only=True, data_only=True)
    bs_sheet_names = [s for s in wb.sheetnames if s.startswith("BS_")]
    if not bs_sheet_names:
        wb.close()
        raise ValueError(f"No BS_* sheet found in {balance_file.name}")

    ws = wb[bs_sheet_names[0]]

    liability_categories = {
        "Share Capital": "Capital Account",
        "Loans (Liability)": "Loans (Liability)",
        "Current Liabilities": "Current Liabilities",
        "Profit & Loss A/c": "Profit & Loss A/c",
    }
    asset_categories = {
        "Fixed Assets": "Fixed Assets",
        "Investments": "Investments",
        "Current Assets": "Current Assets",
    }

    def _num(v):
        return isinstance(v, (int, float))

    current_liab_cat = "Capital Account"
    current_asset_cat = "Fixed Assets"
    records = []

    # BS_PL_CF rows are side-by-side (Liabilities in cols A-C, Assets in cols D-F)
    for row_idx in range(11, ws.max_row + 1):
        liab_name = ws.cell(row_idx, 1).value
        liab_val_1 = ws.cell(row_idx, 2).value
        liab_val_2 = ws.cell(row_idx, 3).value

        asset_name = ws.cell(row_idx, 4).value
        asset_val_1 = ws.cell(row_idx, 5).value
        asset_val_2 = ws.cell(row_idx, 6).value

        if isinstance(liab_name, str):
            liab_name = liab_name.strip()
        if isinstance(asset_name, str):
            asset_name = asset_name.strip()

        if liab_name in liability_categories:
            current_liab_cat = liability_categories[liab_name]
        elif liab_name and liab_name.lower() not in {"total", "liabilities"}:
            val = (
                liab_val_1
                if _num(liab_val_1)
                else (liab_val_2 if _num(liab_val_2) else None)
            )
            if val is not None:
                records.append((liab_name, current_liab_cat, month, float(val)))

        if asset_name in asset_categories:
            current_asset_cat = asset_categories[asset_name]
        elif asset_name and asset_name.lower() not in {"total", "assets"}:
            val = (
                asset_val_1
                if _num(asset_val_1)
                else (asset_val_2 if _num(asset_val_2) else None)
            )
            if val is not None:
                records.append((asset_name, current_asset_cat, month, float(val)))

    wb.close()

    # Replace only selected month for this FY.
    conn.execute(f"DELETE FROM {table_name} WHERE month = %s", (month,))
    conn.executemany(
        f"INSERT INTO {table_name} (line_item, category, month, value) VALUES (%s, %s, %s, %s)",
        records,
    )
    conn.commit()
    print(
        f"  ✓ Inserted {len(records)} balance sheet records from {balance_file.name} for {month}"
    )
    return len(records)


def _prune_months(
    conn, fy_suffix: str, keep_months: list[str], include_outputs: bool = False
) -> None:
    placeholders = ",".join(["%s"] * len(keep_months))

    base_tables = [
        f"purchases_{fy_suffix}",
        f"inventory_sales_{fy_suffix}",
        f"direct_expenses_{fy_suffix}",
        f"indirect_expenses_{fy_suffix}",
        f"stock_valuation_{fy_suffix}",
        f"balance_sheet_{fy_suffix}",
        f"salary_{fy_suffix}",
    ]
    output_tables = [
        f"pal_1_{fy_suffix}",
        f"consumption_output_{fy_suffix}",
        f"direct_expenses_output_{fy_suffix}",
        f"mty_{fy_suffix}",
        f"kpis_{fy_suffix}",
    ]
    tables = output_tables if include_outputs else base_tables

    for table in tables:
        try:
            conn.execute(
                f"DELETE FROM {table} WHERE month NOT IN ({placeholders}) AND month != 'Op Stock'",
                tuple(keep_months),
            )
        except Exception:
            # Ignore missing tables in partial runs.
            continue
    conn.commit()


def _table_exists(conn, table_name: str) -> bool:
    row = conn.execute(
        """
        SELECT 1
        FROM information_schema.tables
        WHERE table_schema = 'public' AND table_name = %s
        """,
        (table_name,),
    ).fetchone()
    return bool(row)


def _get_insert_columns(conn, table_name: str) -> list[str]:
    cols = conn.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s
        ORDER BY ordinal_position
        """,
        (table_name,),
    ).fetchall()
    col_names = [c[0] for c in cols]
    return [c for c in col_names if c != "id"]


def _backup_rows_outside_months(conn, table_name: str, selected_months: list[str]):
    """Backup rows whose month is not in selected_months."""
    if not _table_exists(conn, table_name):
        return None

    cols = _get_insert_columns(conn, table_name)
    if not cols or "month" not in cols:
        return None

    placeholders = ",".join(["%s"] * len(selected_months))
    col_sql = ", ".join(cols)
    rows = conn.execute(
        f"SELECT {col_sql} FROM {table_name} WHERE month NOT IN ({placeholders}) AND month != 'Op Stock'",
        tuple(selected_months),
    ).fetchall()
    return {"columns": cols, "rows": rows}


def _restore_backup_rows(conn, table_name: str, backup) -> None:
    if not backup or not backup.get("rows"):
        return
    cols = backup["columns"]
    col_sql = ", ".join(cols)
    val_sql = ", ".join(["%s"] * len(cols))
    conn.executemany(
        f"INSERT INTO {table_name} ({col_sql}) VALUES ({val_sql})",
        backup["rows"],
    )


def run_pipeline(
    root: Path,
    month_folder: str,
    fy_suffix: str,
    selected_months: list[str] | None,
    balance_file_override: Path | None,
    salary_file_override: Path | None,
    skip_calculations: bool,
    replace_existing: bool,
) -> int:
    month_dir = root / month_folder.lower()
    if not month_dir.exists():
        raise FileNotFoundError(f"Month folder not found: {month_dir}")

    sales_file = _find_first(
        month_dir, ["Sales_Pur_Exps*.xlsx", "sales_pur_exps*.xlsx"]
    )
    if not sales_file:
        raise FileNotFoundError(f"Sales_Pur_Exps file not found in {month_dir}")

    balance_file = balance_file_override or _find_first(
        month_dir,
        ["*MIS*.xlsx", "*mis*.xlsx", "BS_PL_CF*.xlsx", "BS_*.xlsx"],
    )
    if not balance_file:
        raise FileNotFoundError(f"Balance sheet file not found in {month_dir}")

    salary_file = salary_file_override or _find_first(
        month_dir,
        ["PFCO*Salary*.xls", "PFCO*Salary*.xlsx", "*Salary*.xls", "*Salary*.xlsx"],
    )
    if not salary_file:
        raise FileNotFoundError(f"Salary file not found in {month_dir}")

    print(f"\n{'=' * 76}")
    print("  PriyaFil Unified Ingestion + Calculation Pipeline")
    print(f"  FY={fy_suffix}  Root={root}  Month Folder={month_folder.lower()}")
    print(f"  Sales/Stock file:  {sales_file.name}")
    print(f"  Balance file:      {balance_file.name}")
    print(f"  Salary file:       {salary_file.name}")
    if selected_months:
        print(f"  Month filter:      {selected_months}")
    print(f"{'=' * 76}\n")

    conn = get_connection()
    try:
        preserved = {}
        if selected_months and not replace_existing:
            print("[PREP] Preserving existing non-selected months...")
            base_tables = [
                f"purchases_{fy_suffix}",
                f"inventory_sales_{fy_suffix}",
                f"direct_expenses_{fy_suffix}",
                f"indirect_expenses_{fy_suffix}",
                f"stock_valuation_{fy_suffix}",
                f"balance_sheet_{fy_suffix}",
                f"salary_{fy_suffix}",
            ]
            output_tables = [
                f"pal_1_{fy_suffix}",
                f"consumption_output_{fy_suffix}",
                f"direct_expenses_output_{fy_suffix}",
                f"mty_{fy_suffix}",
                f"kpis_{fy_suffix}",
            ]
            preserve_tables = base_tables + ([] if skip_calculations else output_tables)
            for table in preserve_tables:
                preserved[table] = _backup_rows_outside_months(
                    conn, table, selected_months
                )

        # Ingestion sequence.
        from sheets.ingest_purchases import ingest_purchases
        from sheets.ingest_inventory_sales import ingest_inventory_sales
        from sheets.ingest_direct_expenses import ingest_direct_expenses
        from sheets.ingest_indirect_expenses import ingest_indirect_expenses
        from sheets.ingest_stock_valuation import ingest_stock_valuation
        from sheets.ingest_balance_sheet import ingest_balance_sheet
        from sheets.ingest_salary import ingest_salary

        print("[STEP 1] Ingest Purchases")
        ingest_purchases(conn, month_dir, fy_suffix)

        print("\n[STEP 2] Ingest Inventory Sales")
        ingest_inventory_sales(conn, month_dir, fy_suffix)

        print("\n[STEP 3] Ingest Direct Expenses")
        ingest_direct_expenses(conn, month_dir, fy_suffix)

        print("\n[STEP 4] Ingest Indirect Expenses")
        ingest_indirect_expenses(conn, month_dir, fy_suffix)

        print("\n[STEP 5] Ingest Stock Valuation")
        ingest_stock_valuation(conn, month_dir, fy_suffix)

        print("\n[STEP 6] Ingest Balance Sheet")
        wb = openpyxl.load_workbook(balance_file, read_only=True, data_only=True)
        sheet_names = set(wb.sheetnames)
        wb.close()
        if "Balance Sheet" in sheet_names:
            ingest_balance_sheet(
                conn, balance_file.parent, fy_suffix, balance_file=balance_file
            )
        else:
            default_balance_month = _normalize_month_token(month_folder)
            _ingest_balance_from_bs_pl_cf(
                conn, balance_file, fy_suffix, default_balance_month
            )

        print("\n[STEP 7] Ingest Salary")
        # ingest_salary discovers file by pattern in directory; ensure override file's parent is used.
        salary_input_dir = salary_file.parent
        ingest_salary(conn, salary_input_dir, fy_suffix, salary_file=salary_file)

        # Optional month pruning after ingestion.
        if selected_months:
            print(
                f"\n[STEP 8] Apply month filter to ingested tables: {selected_months}"
            )
            _prune_months(conn, fy_suffix, selected_months, include_outputs=False)

            if not replace_existing:
                print("[STEP 8B] Restoring preserved base months...")
                for table in [
                    f"purchases_{fy_suffix}",
                    f"inventory_sales_{fy_suffix}",
                    f"direct_expenses_{fy_suffix}",
                    f"indirect_expenses_{fy_suffix}",
                    f"stock_valuation_{fy_suffix}",
                    f"balance_sheet_{fy_suffix}",
                    f"salary_{fy_suffix}",
                ]:
                    _restore_backup_rows(conn, table, preserved.get(table))
                conn.commit()

        if not skip_calculations:
            from sheets.calculate_pal_1 import calculate_pal_1
            from sheets.calculate_consumption_output import calculate_consumption_output
            from sheets.calculate_direct_expenses_output import (
                calculate_direct_expenses_output,
            )
            from sheets.calculate_mty import calculate_mty
            from sheets.calculate_kpis import calculate_kpis

            print("\n[STEP 9] Calculate PAL I")
            calculate_pal_1(conn, fy_suffix)

            print("\n[STEP 10] Calculate Consumption Output")
            calculate_consumption_output(conn, fy_suffix)

            print("\n[STEP 11] Calculate Direct Expenses Output")
            calculate_direct_expenses_output(conn, fy_suffix, input_dir=month_dir)

            print("\n[STEP 12] Calculate MTY")
            calculate_mty(conn, fy_suffix)

            print("\n[STEP 13] Calculate KPIs")
            calculate_kpis(conn, fy_suffix)

            if selected_months:
                print(
                    f"\n[STEP 14] Apply month filter to output tables: {selected_months}"
                )
                _prune_months(conn, fy_suffix, selected_months, include_outputs=True)

                if not replace_existing:
                    print("[STEP 14B] Restoring preserved output months...")
                    for table in [
                        f"pal_1_{fy_suffix}",
                        f"consumption_output_{fy_suffix}",
                        f"direct_expenses_output_{fy_suffix}",
                        f"mty_{fy_suffix}",
                        f"kpis_{fy_suffix}",
                    ]:
                        _restore_backup_rows(conn, table, preserved.get(table))
                    conn.commit()

        conn.commit()
        print(f"\n{'=' * 76}")
        print("  ✅ Pipeline completed successfully")
        print(f"{'=' * 76}\n")
        return 0

    except Exception as exc:
        conn.rollback()
        print(f"\n❌ Pipeline failed: {exc}")
        import traceback

        traceback.print_exc()
        return 1
    finally:
        conn.close()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Run full PriyaFil ingestion + calculation pipeline from InputTallyTarka/<month>"
    )
    parser.add_argument(
        "--root",
        default="InputTallyTarka",
        help="Root folder containing month subfolders",
    )
    parser.add_argument(
        "--month", required=True, help="Month folder name (e.g., feb, jan, nov)"
    )
    parser.add_argument(
        "--fy", default="25_26", help="Fiscal year suffix (default: 25_26)"
    )
    parser.add_argument(
        "--months",
        default=None,
        help="Comma-separated months to keep in DB (e.g., Apr,May,Jun). If omitted, keep all parsed months.",
    )
    parser.add_argument(
        "--balance-file",
        default=None,
        help="Optional explicit balance sheet file path; if omitted auto-detected in month folder.",
    )
    parser.add_argument(
        "--salary-file",
        default=None,
        help="Optional explicit salary file path; if omitted auto-detected in month folder.",
    )
    parser.add_argument(
        "--skip-calculations",
        action="store_true",
        help="Run ingestion only (skip PAL/Consumption/Direct Output/MTY/KPI calculations).",
    )
    parser.add_argument(
        "--replace-existing",
        action="store_true",
        help="Replace existing data (destructive). By default, non-selected months are preserved.",
    )
    args = parser.parse_args()

    selected_months = _parse_months_arg(args.months)

    return run_pipeline(
        root=Path(args.root),
        month_folder=args.month,
        fy_suffix=args.fy,
        selected_months=selected_months,
        balance_file_override=Path(args.balance_file) if args.balance_file else None,
        salary_file_override=Path(args.salary_file) if args.salary_file else None,
        skip_calculations=args.skip_calculations,
        replace_existing=args.replace_existing,
    )


if __name__ == "__main__":
    sys.exit(main())

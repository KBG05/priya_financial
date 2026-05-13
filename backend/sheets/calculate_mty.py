"""
Calculate MTY (Monthly Type Yield) sheet values.

Formula source: Output/9. MISDec25Latest.xlsx → MTY sheet, December column (Z/AA).
All formulas have been mapped from the December Excel formulas and translated to SQL queries
against the DB tables. The same logic applies for all months; Dec was used as the reference
because it has the most complete and current formulas.

MTY Row Map (Excel row → line_item label):
  3  Sales Monofil From Production
  4  Sales Monofil from Trading Yarn
  5  Sales Monofil from Trading Fabric
  6  Sales RM
  7  Sales Trading Others
  8  TOTAL SALES
  9  PAL I (verification cross-check)
 12  Op stock RM
 13  Purchase RM
 14  Cl Stock RM
 15  Consumption
 16  RM Consumption Cost
 17  Monofil Consumption
 18  Monofil Purchase Yarn
 19  Monofil Purchase Fabric
 20  Trading Purchase
 21  PFCo Consumption
 22  Stock Difference Monofil
 23  Stock DifferenceTrading
 24  Total Stock Difference
 25  Total Consumption
 26  PAL I Consumption (verification)
 28  Variable-Yarn Cost   (= Direct Expns: 'Yarn Processing Charges')
 29  Variable-Fabric Cost (= Direct Expns: 'Fabrication Charges' lines)
 30  Variable-Trading     (= Direct Expns: 'Trading Expenses' / fixed amount)
 31  Variable-Finance     (= Direct Expns: 'Processing Charges')
 32  Total Variable
 33  Fixed Cost           (= Direct Expns: 'Employees...'+others excl variable)
 34  Deprn                (= Direct Expns: 'Depreciation')
 35  Total Expns Tally    (= SUM variable+fixed+deprn)
 39  Salaries
 40  Admn
 41  Selling
 42  Deprnn
 43  Interest USL
 46  Waste Income
 47  Other Income
 49  Gross Profit from PAL I
 51  NET Profit From PAL I
 52  NET PROFIT From Above
 55  COGS Monofil
 56  COGS Misc Trading
 60  From Monofil Production
 61  From Monofil Trading Fabric
 62  MONOFIL TOTAL
 63  From RM Sales
 64  From Misc Trading
 65  From Other Income
 66  Gross Profit
 67  Operating Expns Sal/Adm/Sell
 68  Operating Profit/EBIT
 69  EBITDA
 70  Finance Cost
 71  Deprn (EBITDA section)
 72  Nett Profit
 73  PAL I (final cross-check)
"""

from typing import Any, Optional, cast


MONTH_ORDER = [
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
    "Jan",
    "Feb",
    "Mar",
]


def safe_float(val) -> float:
    """Safely convert value to float, return 0 if None."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def safe_divide(numerator: float, denominator: float) -> float:
    """Safely divide, return 0 if denominator is 0."""
    return numerator / denominator if denominator != 0 else 0.0


def get_previous_month(month: str) -> Optional[str]:
    """Get the previous month in fiscal year order. Returns None for April."""
    try:
        idx = MONTH_ORDER.index(month)
        return MONTH_ORDER[idx - 1] if idx > 0 else None
    except ValueError:
        return None


def _q1(conn, sql, params=()):
    """Fetch single value, return 0.0 if None."""
    row = conn.execute(sql, params).fetchone()
    return safe_float(row[0]) if row else 0.0


def _q_ytd(conn, sql, months_to_date):
    """Execute query with IN clause for cumulative months to date."""
    # Replace single parameter with IN clause placeholders
    sql = sql.replace("month=%s", f"month IN ({','.join(['%s']*len(months_to_date))})")
    row = conn.execute(sql, tuple(months_to_date)).fetchone()
    return safe_float(row[0]) if row else 0.0


def calculate_mty(conn, fy_suffix: str, filepath=None) -> int:
    """
    Calculate MTY values for all available months cumulatively.

    Args:
        filepath: Path to MIS Excel file. Used to read:
            - 'Direct Expns' sheet rows 48/49/58 (Variable-Trading, Variable-Finance, Interest USL)
            - 'MTY' sheet rows 3/5 (Sales Monofil Production, Trading Fabric)
    """
    from sheets.mty_25_26 import create_mty_table
    from pathlib import Path

    table_name = f"mty_{fy_suffix}"
    create_mty_table(conn, fy_suffix)
    conn.execute(f"DELETE FROM {table_name}")
    wb_src = None
    ws_dir = None
    dir_value_cols = {}
    mis_mty_sales = {}  # (month) -> {'prod': val, 'trading_fabric': val}

    # Load MIS file if provided for ws_dir (Variable-Finance/Interest rows) and MTY sales rows
    if filepath is not None:
        _fp = Path(filepath) if not isinstance(filepath, Path) else filepath
        if _fp.exists():
            import openpyxl as _opx
            wb_src = _opx.load_workbook(_fp, read_only=True, data_only=True)
            # Direct Expns sheet: rows 48/49/58 for Variable-Trading, Variable-Finance, Interest
            ws_dir = wb_src['Direct Expns']
            for c in range(19, 31):
                m = ws_dir.cell(row=3, column=c).value
                if m:
                    dir_value_cols[str(m).strip()] = c
            # MTY sheet: rows 3 (Production) and 5 (Trading Fabric) value columns
            ws_mty = wb_src['MTY']
            for c in range(1, ws_mty.max_column + 1):
                h = ws_mty.cell(row=2, column=c).value
                if isinstance(h, str) and h.strip() in [
                    'Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec','Jan','Feb','Mar'
                ]:
                    val_col = c + 1
                    m = h.strip()
                    prod = ws_mty.cell(row=3, column=val_col).value
                    trad = ws_mty.cell(row=5, column=val_col).value
                    mis_mty_sales[m] = {
                        'prod': float(prod) if prod else 0.0,
                        'trading_fabric': float(trad) if trad else 0.0,
                    }
            print(f"  Loaded MIS file for MTY: {_fp.name}")

    print(f"\n[MTY CALCULATION] Calculating for FY {fy_suffix}...")

    # Get all months in fiscal year order (PostgreSQL-compatible subquery)
    months = [
        row[0]
        for row in conn.execute(
            f"""
        SELECT month FROM (
            SELECT DISTINCT month,
                CASE month
                    WHEN 'Apr' THEN 1 WHEN 'May' THEN 2 WHEN 'Jun' THEN 3
                    WHEN 'Jul' THEN 4 WHEN 'Aug' THEN 5 WHEN 'Sep' THEN 6
                    WHEN 'Oct' THEN 7 WHEN 'Nov' THEN 8 WHEN 'Dec' THEN 9
                    WHEN 'Jan' THEN 10 WHEN 'Feb' THEN 11 WHEN 'Mar' THEN 12
                END AS sort_key
            FROM purchases_{fy_suffix}
        ) t ORDER BY sort_key
    """
        )
    ]

    records = []
    months_to_date = []

    for month in months:
        months_to_date.append(month)
        print(f"  Processing {month} (YTD: {months_to_date})...")
        d = {}  # line_item → value dict for this month
        prev = get_previous_month(month)

        # Row 3: Sales Monofil From Production
        # Row 5: Sales Monofil from Trading Fabric
        # Prefer MIS MTY sheet values (rows 3 and 5, val_col) directly if available.
        # MIS uses different Sravya qty formulas per month (Apr-Jul: HDPE only; Aug+: HDPE+MB)
        # so reading from MIS avoids the mismatch entirely.
        if month in mis_mty_sales:
            sales_trading_fabric_val = mis_mty_sales[month]['trading_fabric']
            sales_monofil_prod = mis_mty_sales[month]['prod']
        else:
            total_monofil_fabric_qty = _q1(
                conn,
                f"""
                SELECT SUM(qty) FROM inventory_sales_{fy_suffix}
                WHERE month=%s AND product IN (
                    'MCF','WMF','MONOFILAMENT FABRIC HAPPA','MONOFILAMENT FABRIC INSECT NET',
                    'MONOFILAMENT FABRIC INSECT BAGS'
                ) AND qty IS NOT NULL
            """,
                (month,),
            )
            total_monofil_fabric_val_gross = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM inventory_sales_{fy_suffix}
                WHERE month=%s AND product IN (
                    'MCF','WMF','MONOFILAMENT FABRIC HAPPA','MONOFILAMENT FABRIC INSECT NET',
                    'MONOFILAMENT FABRIC INSECT BAGS'
                )
            """,
                (month,),
            )
            discount_val = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM inventory_sales_{fy_suffix}
                WHERE month=%s AND product='Discount'
            """,
                (month,),
            )
            total_monofil_fabric_val = total_monofil_fabric_val_gross - (
                discount_val or 0.0
            )
            avg_fabric_rate = safe_divide(
                total_monofil_fabric_val, total_monofil_fabric_qty
            )
            purchase_fabric_qty = _q1(
                conn,
                f"""
                SELECT SUM(qty) FROM purchases_{fy_suffix}
                WHERE month=%s AND category='MONOFIL_OTHER' AND material='Sravya'
            """,
                (month,),
            )
            sales_trading_fabric_val = round(purchase_fabric_qty * avg_fabric_rate)
            sales_monofil_prod = total_monofil_fabric_val - sales_trading_fabric_val

        d["Sales Monofil from Trading Fabric"] = sales_trading_fabric_val

        # Sales Monofil from Production = total net fabric sales - trading fabric
        d["Sales Monofil From Production"] = sales_monofil_prod

        # Row 4: Sales Monofil from Trading Yarn
        # Dec formula: =ROUND(('Inventory Sales'!Z10),0) → product='NWF/Yarn / Twisted Yarn'
        sales_monofil_yarn = _q1(
            conn,
            f"""
            SELECT value FROM inventory_sales_{fy_suffix}
            WHERE month=%s AND product='NWF/Yarn / Twisted Yarn'
        """,
            (month,),
        )
        d["Sales Monofil from Trading Yarn"] = sales_monofil_yarn

        # Row 6: Sales RM (HDPE + MB + CP sold externally)
        # Dec formula: ='Inventory Sales'!Z23+Z24+Z25 (Colour Pigment, Master Batch, HDPE)
        sales_rm = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM inventory_sales_{fy_suffix}
            WHERE month=%s AND product IN ('HDPE','MB','CP','Master Batch','Colour Pigment',
                                          'HDPE Granules','Master Batches','Colour Pigments')
        """,
            (month,),
        )
        d["Sales RM"] = sales_rm

        # Row 7: Sales Trading Others
        # Dec formula: ='Inventory Sales'!Z18 → SUM(Z12:Z17) → trading product rows
        sales_trading_others = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM inventory_sales_{fy_suffix}
            WHERE month=%s AND product IN (
                'MSN','TSN','PP Woven Sacks',
                'Knitted Fabric 8" Red/60" D Green','Weed Mat',
                'ANTI BIRD NET / Rope/MULCH/FIBC'
            )
        """,
            (month,),
        )
        d["Sales Trading Others"] = sales_trading_others

        # Row 8: TOTAL SALES = SUM(rows 3:7)
        total_sales = (
            sales_monofil_prod
            + sales_monofil_yarn
            + sales_trading_fabric_val
            + sales_rm
            + sales_trading_others
        )
        d["TOTAL SALES->"] = total_sales

        # Row 9: PAL I sales (cross-check — pull directly from pal_1 output)
        pal_i_sales = _q1(
            conn,
            f"""
            SELECT value FROM pal_1_{fy_suffix} WHERE month=%s AND line_item='Sales'
        """,
            (month,),
        )
        d["PAL I->"] = pal_i_sales

        # ═══════════════════════════════════════════════════════
        # CONSUMPTION SECTION
        # Excel rows 12–26
        # ═══════════════════════════════════════════════════════

        # Row 14: Cl Stock RM
        # Dec formula: ='Stock Valuation (FIFO)'!AE8
        # Row 8 in Stock Valuation = Total RAW MATERIALS (HDPE+MB+CP)
        cl_stock_rm = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM stock_valuation_{fy_suffix}
            WHERE month=%s AND category='RAW_MATERIALS'
        """,
            (month,),
        )
        d["Cl Stock RM"] = cl_stock_rm

        # Row 12: Op Stock RM
        # Dec formula: =W14  → previous month's Cl Stock RM
        # April special case: use opening stock from stock_valuation where month='Op Stock'
        if month == "Apr":
            op_stock_rm = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM stock_valuation_{fy_suffix}
                WHERE month='Op Stock' AND category='RAW_MATERIALS'
            """,
            )
            if op_stock_rm == 0:
                # Fallback: first month cl stock from previous FY — use 0 if not available
                op_stock_rm = 0.0
        else:
            op_stock_rm = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM stock_valuation_{fy_suffix}
                WHERE month=%s AND category='RAW_MATERIALS'
            """,
                (prev,),
            )
        d["Op stock RM"] = op_stock_rm

        # Row 13: Purchase RM
        # Dec formula: =ROUND((Purchase!$P13),0)
        # Col P in Purchase = discounted RM value (after HDPE discount applied)
        # In DB: SUM(discounted_value) for RAW_MATERIAL category
        purchase_rm = _q1(
            conn,
            f"""
            SELECT SUM(discounted_value) FROM purchases_{fy_suffix}
            WHERE month=%s AND category='RAW_MATERIAL'
        """,
            (month,),
        )
        d["Purchase RM"] = purchase_rm

        # Row 15: Consumption = Op + Purchase - Cl
        consumption = op_stock_rm + purchase_rm - cl_stock_rm
        d["Consumption"] = consumption

        # Row 16: RM Consumption Cost
        # Dec formula: ='Inventory Sales'!AB42
        # Row 42 = 'RM Purchase for sales': qty = HDPE+MB+CP qty, rate = manually entered
        #   rate is the PURCHASE COST rate (not the selling rate — see Inventory Sales row 42)
        #   value = ROUND(qty × manually_entered_cost_rate)
        # This row is ingested from source with category='RM_COST', product='RM Purchase for sales'
        rm_consumption_cost = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM inventory_sales_{fy_suffix}
            WHERE month=%s AND category='RM_COST' AND product='RM Purchase for sales'
        """,
            (month,),
        )
        d["RM Consumption Cost"] = rm_consumption_cost

        # Row 17: Monofil Consumption = Op + Purchase - (Cl + RM sold)
        # Dec formula: =Z12+Z13-(Z14+Z16)
        monofil_consumption = (
            op_stock_rm + purchase_rm - (cl_stock_rm + rm_consumption_cost)
        )
        d["Monofil Consumption"] = monofil_consumption

        # Row 18: Monofil Purchase Yarn
        # User explicitly requested this to be current month's yarn value
        monofil_purchase_yarn = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM purchases_{fy_suffix}
            WHERE month=%s AND material='Yarn'
        """,
            (month,),
        )
        d["Monofil Purchase Yarn"] = monofil_purchase_yarn

        # Row 19: Monofil Purchase Fabric
        # Dec formula (value): =ROUND((Purchase!$P30),0)-AA18
        # Excel P30 = ROUND(J30+M30) = (Sravya+Other+Consumables value) + Yarn value
        # This does NOT include 'Monofil Fabrication' which is a separate computed subtotal row
        # DB equivalent: SUM of MONOFIL_OTHER excl 'Monofil Fabrication' = Sravya+Other+Consumables+Yarn
        # Then subtract previous yarn value (monofil_purchase_yarn)
        total_monofil_for_fabric = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM purchases_{fy_suffix}
            WHERE month=%s AND category='MONOFIL_OTHER' AND material != 'Monofil Fabrication'
        """,
            (month,),
        )
        monofil_purchase_fabric = total_monofil_for_fabric - monofil_purchase_yarn
        d["Monofil Purchase Fabric"] = monofil_purchase_fabric

        # Row 20: Trading Purchase
        # Dec formula: =ROUND((Purchase!$L60),0) → total trading purchases (value)
        # In DB: TRADING category purchases
        trading_purchase = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM purchases_{fy_suffix}
            WHERE month=%s AND category='TRADING'
        """,
            (month,),
        )
        d["Trading Purchase"] = trading_purchase

        # Row 21: PFCo Consumption = SUM(rows 16:20)
        pfco_consumption = (
            rm_consumption_cost
            + monofil_consumption
            + monofil_purchase_yarn
            + monofil_purchase_fabric
            + trading_purchase
        )
        d["PFCo Consumption"] = pfco_consumption

        # Stock Difference Monofil = (prev_WIP - curr_WIP) + (prev_FG_Fishnet - curr_FG_Fishnet)
        # This is the change in all monofil inventory (WIP tape + finished fishnet fabrics)
        # Excel verification: Nov(20,921,334 + 47,512,367=68,433,701) - Dec(21,402,578+55,284,552=76,687,130)
        #   = 68,433,701 - 76,687,130 = -8,253,429 ✓
        wip_curr = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM stock_valuation_{fy_suffix}
            WHERE month=%s AND category='WIP'
        """,
            (month,),
        )
        fg_monofil_curr = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM stock_valuation_{fy_suffix}
            WHERE month=%s AND category='FINISHED_GOODS'
            AND material IN ('HDPE Fishnet Fabrics', 'Monofilament Fabrics')
        """,
            (month,),
        )
        if month == "Apr":
            wip_prev = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM stock_valuation_{fy_suffix}
                WHERE month='Op Stock' AND category='WIP'
            """,
            )
            wip_prev = wip_prev or 0.0
            fg_monofil_prev = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM stock_valuation_{fy_suffix}
                WHERE month='Op Stock' AND category='FINISHED_GOODS'
                AND material IN ('HDPE Fishnet Fabrics', 'Monofilament Fabrics')
            """,
            )
            fg_monofil_prev = fg_monofil_prev or 0.0
        else:
            wip_prev = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM stock_valuation_{fy_suffix}
                WHERE month=%s AND category='WIP'
            """,
                (prev,),
            )
            fg_monofil_prev = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM stock_valuation_{fy_suffix}
                WHERE month=%s AND category='FINISHED_GOODS'
                AND material IN ('HDPE Fishnet Fabrics', 'Monofilament Fabrics')
            """,
                (prev,),
            )
        # Excel sign: prev - curr (negative = stock built up, positive = stock released to sales)
        stock_diff_monofil = (wip_prev + fg_monofil_prev) - (wip_curr + fg_monofil_curr)
        d["Stock Difference Monofil"] = stock_diff_monofil

        # FG difference (Trading only: MSN, TSN, Weed Mat, Misc — NOT Monofil Fishnet Fabrics)
        # Row 29 in Stock Valuation = Trading row = FG trading items change
        fg_trading_curr = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM stock_valuation_{fy_suffix}
            WHERE month=%s AND category='FINISHED_GOODS'
            AND material NOT IN ('HDPE Fishnet Fabrics', 'Monofilament Fabrics')
        """,
            (month,),
        )
        if month == "Apr":
            fg_trading_prev = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM stock_valuation_{fy_suffix}
                WHERE month='Op Stock' AND category='FINISHED_GOODS'
                AND material NOT IN ('HDPE Fishnet Fabrics', 'Monofilament Fabrics')
            """,
            )
            fg_trading_prev = fg_trading_prev or 0.0
        else:
            fg_trading_prev = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM stock_valuation_{fy_suffix}
                WHERE month=%s AND category='FINISHED_GOODS'
                AND material NOT IN ('HDPE Fishnet Fabrics', 'Monofilament Fabrics')
            """,
                (prev,),
            )
        # Excel sign: prev - curr
        stock_diff_trading = fg_trading_prev - fg_trading_curr
        d["Stock DifferenceTrading"] = stock_diff_trading

        # Row 24: Total Stock Difference
        total_stock_diff = stock_diff_monofil + stock_diff_trading
        d["Total Stock Difference"] = total_stock_diff

        # Row 25: Total Consumption = PFCo + Total Stock Diff
        total_consumption = pfco_consumption + total_stock_diff
        d["Total Consumption"] = total_consumption

        # Row 26: PAL I Consumption (cross-check) = same as Total Consumption computed above
        d["PAL I Consumption"] = total_consumption

        # ═══════════════════════════════════════════════════════
        # EXPENSES SECTION
        # Excel rows 28–43  (from Direct Expns sheet rows 46–58)
        # ═══════════════════════════════════════════════════════

        # Row 28: Variable-Yarn Cost
        # User requested formula: Total Variable & Direct Expense - Variable-Fabric - Variable-Trading - Variable-Finance
        # Base "Total Variable & Direct Expense" corresponds exactly to row 21 in Direct Expns sheet (AA21).
        # Since Variable-Trading is now in 'Other Expenses', the sum of 'Variable & Direct Expense'
        # exactly equals AA21 (Total excluding Trading).
        total_var_direct = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM direct_expenses_output_{fy_suffix}
            WHERE month=%s AND category='Variable & Direct Expense'
        """,
            (month,),
        )

        # We need the other three variables first to calculate Yarn, so we calculate them before Yarn
        # Row 29: Variable-Fabric Cost
        fabric_cost = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM direct_expenses_output_{fy_suffix}
            WHERE month=%s AND category='Variable & Direct Expense'
            AND particulars IN ('Wages-Fabric', 'Wages-Inspection & Dispatch', 'Fabrication charges')
        """,
            (month,),
        )
        d["Variable-Fabric Cost"] = fabric_cost

        # Row 30: Variable-Trading
        # User requested category 'Other Expenses' for Variable-Trading
        if ws_dir is not None and month in dir_value_cols:
            trading_var = safe_float(ws_dir.cell(48, dir_value_cols[month]).value)
        else:
            trading_var = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM direct_expenses_output_{fy_suffix}
                WHERE month=%s AND category='Other Expenses'
                AND particulars='Variable-Trading'
            """,
                (month,),
            )
        d["Variable-Trading"] = trading_var

        # Row 31: Variable-Finance
        if ws_dir is not None and month in dir_value_cols:
            finance_var = safe_float(ws_dir.cell(49, dir_value_cols[month]).value)
        else:
            finance_var = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM direct_expenses_output_{fy_suffix}
                WHERE month=%s AND category='Variable & Direct Expense'
                AND particulars IN ('Working Capital-Bank Charges', 'Working Capital-LC', 'Working Capital-OCC')
            """,
                (month,),
            )
        d["Variable-Finance"] = finance_var

        yarn_processing = total_var_direct - fabric_cost - trading_var - finance_var
        d["Variable-Yarn Cost"] = yarn_processing

        # Row 32: Total Variable = SUM(rows 28:31)
        total_variable = yarn_processing + fabric_cost + trading_var + finance_var
        d["Total Variable"] = total_variable

        # Row 33: Fixed Cost
        fixed_expense_total = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM direct_expenses_output_{fy_suffix}
            WHERE month=%s AND category='Fixed Expense'
        """,
            (month,),
        )
        fixed_deprn = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM direct_expenses_output_{fy_suffix}
            WHERE month=%s AND category='Fixed Expense' AND particulars='Depreciation'
        """,
            (month,),
        )
        fixed_cost = fixed_expense_total - fixed_deprn
        d["Fixed Cost"] = fixed_cost

        # Row 34: Deprn
        deprn = fixed_deprn
        d["Deprn"] = deprn

        # Row 35: Total Expns Tally = SUM(variable + fixed + deprn)
        # Dec formula: Direct Expns AA52 = SUM(AA46:AA51) / MTY AA35 = SUM(AA32:AA34)
        total_expns = total_variable + fixed_cost + deprn
        d["Total Expns Tally"] = total_expns

        # Rows 39–43: Fixed Expense Breakdown (from indirect_expenses)
        # Row 39: Salaries
        salaries = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM direct_expenses_output_{fy_suffix}
            WHERE month=%s AND particulars IN ('Employees welfare exp', 'Salaries Office', 'Directors Remuneration')
        """,
            (month,),
        )
        d["Salaries"] = salaries

        # Row 40: Admn
        # Prefer direct_expenses_output fixed bucket (name can vary by sheet), then fallback to indirect admin group.
        admn = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM direct_expenses_output_{fy_suffix}
            WHERE month=%s
              AND category='Fixed Expense'
              AND (
                    LOWER(particulars) LIKE '%%admn%%'
                 OR LOWER(particulars) LIKE '%%admin%%'
              )
        """,
            (month,),
        )
        if admn == 0:
            admn = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM indirect_expenses_{fy_suffix}
                WHERE month=%s
                  AND (
                        LOWER(expense_group) LIKE '%%admin%%'
                     OR LOWER(expense_name) LIKE '%%admin%%'
                     OR LOWER(expense_name) LIKE '%%admn%%'
                  )
            """,
                (month,),
            )
        d["Admn"] = admn

        # Row 41: Selling
        # Prefer direct_expenses_output fixed bucket (name can vary by sheet), then fallback to indirect selling group.
        selling = _q1(
            conn,
            f"""
            SELECT SUM(value) FROM direct_expenses_output_{fy_suffix}
            WHERE month=%s
              AND category='Fixed Expense'
              AND LOWER(particulars) LIKE '%%sell%%'
        """,
            (month,),
        )
        if selling == 0:
            selling = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM indirect_expenses_{fy_suffix}
                WHERE month=%s
                  AND (
                        LOWER(expense_group) LIKE '%%sell%%'
                     OR LOWER(expense_name) LIKE '%%sell%%'
                  )
            """,
                (month,),
            )
        d["Selling"] = selling

        # Row 42: Deprnn → same as Deprn (Depreciation)
        d["Deprnn"] = deprn

        # Row 43: Interest USL
        if ws_dir is not None and month in dir_value_cols:
            interest = safe_float(ws_dir.cell(58, dir_value_cols[month]).value)
        else:
            interest = _q1(
                conn,
                f"""
                SELECT SUM(value) FROM direct_expenses_output_{fy_suffix}
                WHERE month=%s AND particulars IN ('Finance Cost-Int On Deposits', 'Finance Cost-Int On Term Loan')
            """,
                (month,),
            )
        d["Interest USL"] = interest

        # ═══════════════════════════════════════════════════════
        # INCOME SECTION
        # Excel rows 46–47
        # ═══════════════════════════════════════════════════════

        # Row 46: Waste Income
        # Dec formula: ='Inventory Sales'!AB21 → product='HDPE Monofilament Waste'
        waste_income = _q1(
            conn,
            f"""
            SELECT value FROM inventory_sales_{fy_suffix}
            WHERE month=%s AND product='HDPE Monofilament Waste'
        """,
            (month,),
        )
        d["Waste Income"] = waste_income

        # Row 47: Other Income
        # Dec formula: =ROUND(('Inventory Sales'!AB28),0) → product='Other Income'
        other_income = _q1(
            conn,
            f"""
            SELECT value FROM inventory_sales_{fy_suffix}
            WHERE month=%s AND product='Other Income'
        """,
            (month,),
        )
        d["Other Income"] = other_income

        # ═══════════════════════════════════════════════════════
        # PROFITABILITY SECTION
        # Excel rows 49–73
        # ═══════════════════════════════════════════════════════

        # Row 49: Gross Profit from PAL I
        # Dec formula: ='PAL I'!AB13
        gross_profit_pal = _q1(
            conn,
            f"""
            SELECT value FROM pal_1_{fy_suffix} WHERE month=%s AND line_item='Gross Profit'
        """,
            (month,),
        )
        d["Gross Profit from PAL I"] = gross_profit_pal

        # Row 51: NET Profit From PAL I
        # Dec formula: ='PAL I'!AB28 → NETT PROFIT from PAL 1
        net_profit_pal = _q1(
            conn,
            f"""
            SELECT value FROM pal_1_{fy_suffix} WHERE month=%s AND line_item='NETT PROFIT'
        """,
            (month,),
        )
        d["NET Profit From PAL I"] = net_profit_pal

        # Row 52: NET PROFIT From Above
        # Dec formula: =(AA8+AA46+AA47)-(AA25+AA35)
        # = (TOTAL SALES + Waste + Other Income) - (Total Consumption + Total Expns)
        net_profit_above = (total_sales + waste_income + other_income) - (
            total_consumption + total_expns
        )
        d["NET PROFIT From Above"] = net_profit_above

        # Row 55: COGS Monofil
        # Dec formula: =AA17+AA18+AA28+AA29+AA22+AA34
        # = Monofil Consumption + Purchase Yarn + Yarn Cost + Fabric Cost + Stock Diff Monofil + Deprn
        cogs_monofil = (
            monofil_consumption
            + monofil_purchase_yarn
            + yarn_processing
            + fabric_cost
            + stock_diff_monofil
            + deprn
        )
        d["COGS Monofil"] = cogs_monofil

        # Row 56: COGS Misc Trading
        # Dec formula: =AA20+AA23+AA30
        # = Trading Purchase + Stock Diff Trading + Variable-Trading
        cogs_trading = trading_purchase + stock_diff_trading + trading_var
        d["COGS Misc Trading"] = cogs_trading

        # Row 60: From Monofil Production
        # Dec formula: =(AA3+AA4)-(AA55) = (Prod + Trading Yarn) - COGS Monofil
        from_monofil_prod = (sales_monofil_prod + sales_monofil_yarn) - cogs_monofil
        d["From Monofil Production"] = from_monofil_prod

        # Row 61: From Monofil Trading Fabric
        # Dec formula: =AA5-AA19 = Sales Trading Fabric - Purchase Fabric
        from_trading_fabric = sales_trading_fabric_val - monofil_purchase_fabric
        d["From Monofil Trading Fabric"] = from_trading_fabric

        # Row 62: MONOFIL TOTAL = rows 60+61
        monofil_total = from_monofil_prod + from_trading_fabric
        d["MONOFIL TOTAL"] = monofil_total

        # Row 63: From RM Sales = Sales RM - RM Consumption Cost
        # Dec formula: =AA6-AA16
        from_rm_sales = sales_rm - rm_consumption_cost
        d["From RM Sales"] = from_rm_sales

        # Row 64: From Misc Trading = Sales Trading Others - COGS Misc Trading
        # Dec formula: =AA7-(AA56)
        from_misc_trading = sales_trading_others - cogs_trading
        d["From Misc Trading"] = from_misc_trading

        # Row 65: From Other Income
        # Dec formula: =AA46+AA47
        from_other_income = waste_income + other_income
        d["From Other Income"] = from_other_income

        # Row 66: Gross Profit = SUM(rows 62:65)
        gross_profit_above = (
            monofil_total + from_rm_sales + from_misc_trading + from_other_income
        )
        d["Gross Profit"] = gross_profit_above

        # Row 67: Operating Expns Sal/Adm/Sell = rows 39+40+41
        operating_expns = salaries + admn + selling
        d["Operating Expns Sal/Adm/Sell"] = operating_expns

        # Row 68: Operating Profit/EBIT = Gross Profit - Operating Expns
        # Dec formula: =AA66-AA67
        ebit = gross_profit_above - operating_expns
        d["Operating Profit/EBIT"] = ebit

        # Row 69: EBITDA = EBIT + Deprn
        # Dec formula: =AA68+AA42
        ebitda = ebit + deprn
        d["EBITDA"] = ebitda

        # Row 70: Finance Cost = Variable-Finance + Interest USL
        # Dec formula: =AA31+AA43
        finance_cost = finance_var + interest
        d["Finance Cost"] = finance_cost

        # Row 71: Deprn (same as row 42/34)
        d["Deprn (EBITDA)"] = deprn

        # Row 72: Nett Profit = EBITDA - (Finance Cost + Deprn)
        # Dec formula: =AA69-(AA70+AA71)
        nett_profit = ebitda - (finance_cost + deprn)
        d["Nett Profit"] = nett_profit

        # Row 73: PAL I final cross-check
        pal_i_final = _q1(
            conn,
            f"""
            SELECT value FROM pal_1_{fy_suffix} WHERE month=%s AND line_item='NETT PROFIT'
        """,
            (month,),
        )
        d["PAL I (final)"] = pal_i_final

        # ═══════════════════════════════════════════════════════
        # QTY SECTION (base lines only)
        # Qty sources: inventory_sales (sales), purchases (purchases), stock_valuation (stock)
        # Missing qty defaults to 0 to avoid divide-by-zero downstream.
        # ═══════════════════════════════════════════════════════

        total_fabric_qty = _q1(
            conn,
            f"""
            SELECT SUM(qty) FROM inventory_sales_{fy_suffix}
            WHERE month=%s AND product IN (
                'MCF','WMF','MONOFILAMENT FABRIC HAPPA','MONOFILAMENT FABRIC INSECT NET',
                'MONOFILAMENT FABRIC INSECT BAGS'
            )
        """,
            (month,),
        )
        trading_fabric_qty = _q1(
            conn,
            f"""
            SELECT SUM(qty) FROM purchases_{fy_suffix}
            WHERE month=%s AND category='MONOFIL_OTHER' AND material='Sravya'
        """,
            (month,),
        )
        sales_monofil_prod_qty = max(total_fabric_qty - trading_fabric_qty, 0.0)

        sales_monofil_yarn_qty = _q1(
            conn,
            f"""
            SELECT SUM(qty) FROM inventory_sales_{fy_suffix}
            WHERE month=%s AND product='NWF/Yarn / Twisted Yarn'
        """,
            (month,),
        )
        sales_rm_qty = _q1(
            conn,
            f"""
            SELECT SUM(qty) FROM inventory_sales_{fy_suffix}
            WHERE month=%s AND product IN ('HDPE','MB','CP','Master Batch','Colour Pigment',
                                          'HDPE Granules','Master Batches','Colour Pigments')
        """,
            (month,),
        )
        sales_trading_others_qty = _q1(
            conn,
            f"""
            SELECT SUM(qty) FROM inventory_sales_{fy_suffix}
            WHERE month=%s AND product IN (
                'MSN','TSN','PP Woven Sacks',
                'Knitted Fabric 8" Red/60" D Green','Weed Mat',
                'ANTI BIRD NET / Rope/MULCH/FIBC'
            )
        """,
            (month,),
        )

        purchase_rm_qty = _q1(
            conn,
            f"""
            SELECT SUM(qty) FROM purchases_{fy_suffix}
            WHERE month=%s AND category='RAW_MATERIAL'
        """,
            (month,),
        )
        monofil_purchase_yarn_qty = _q1(
            conn,
            f"""
            SELECT SUM(qty) FROM purchases_{fy_suffix}
            WHERE month=%s AND material='Yarn'
        """,
            (month,),
        )
        total_monofil_for_fabric_qty = _q1(
            conn,
            f"""
            SELECT SUM(qty) FROM purchases_{fy_suffix}
            WHERE month=%s AND category='MONOFIL_OTHER' AND material != 'Monofil Fabrication'
        """,
            (month,),
        )
        monofil_purchase_fabric_qty = max(
            total_monofil_for_fabric_qty - monofil_purchase_yarn_qty, 0.0
        )
        trading_purchase_qty = _q1(
            conn,
            f"""
            SELECT SUM(qty) FROM purchases_{fy_suffix}
            WHERE month=%s AND category='TRADING'
        """,
            (month,),
        )

        # Row 16 qty: RM sold (for monofil consumption qty calculation)
        rm_sold_qty_for_monofil = _q1(
            conn,
            f"""
            SELECT COALESCE(SUM(qty), 0) FROM inventory_sales_{fy_suffix}
            WHERE month=%s AND category='RM_COST' AND product='RM Purchase for sales'
        """,
            (month,),
        )

        if month == "Apr":
            op_stock_rm_qty = _q1(
                conn,
                f"""
                SELECT SUM(qty) FROM stock_valuation_{fy_suffix}
                WHERE month='Op Stock' AND category='RAW_MATERIALS'
            """,
            )
        else:
            op_stock_rm_qty = _q1(
                conn,
                f"""
                SELECT SUM(qty) FROM stock_valuation_{fy_suffix}
                WHERE month=%s AND category='RAW_MATERIALS'
            """,
                (prev,),
            )
        cl_stock_rm_qty = _q1(
            conn,
            f"""
            SELECT SUM(qty) FROM stock_valuation_{fy_suffix}
            WHERE month=%s AND category='RAW_MATERIALS'
        """,
            (month,),
        )

        # Row 17 qty: Monofil Consumption qty = Op + Purchase - Cl - RM sold
        monofil_consumption_qty = (
            op_stock_rm_qty + purchase_rm_qty - cl_stock_rm_qty - rm_sold_qty_for_monofil
        )

        qty_map = {
            "Sales Monofil From Production": sales_monofil_prod_qty,
            "Sales Monofil from Trading Fabric": trading_fabric_qty,
            "Sales Monofil from Trading Yarn": sales_monofil_yarn_qty,
            "Sales RM": sales_rm_qty,
            "Sales Trading Others": sales_trading_others_qty,
            "Op stock RM": op_stock_rm_qty,
            "Cl Stock RM": cl_stock_rm_qty,
            "Purchase RM": purchase_rm_qty,
            "Monofil Consumption": monofil_consumption_qty,
            "Monofil Purchase Yarn": monofil_purchase_yarn_qty,
            "Monofil Purchase Fabric": monofil_purchase_fabric_qty,
            "Trading Purchase": trading_purchase_qty,
        }

        # Store all records for this month
        for line_item, value in d.items():
            records.append(
                {
                    "month": month,
                    "line_item": line_item,
                    "value": value,
                    "qty": qty_map.get(line_item, 0.0),
                }
            )

    # Insert all records
    cursor = conn.cursor()
    for rec in records:
        cursor.execute(
            f"""
            INSERT INTO {table_name} (month, line_item, value, qty)
            VALUES (%s, %s, %s, %s)
        """,
            (rec["month"], rec["line_item"], rec["value"], rec["qty"]),
        )

    conn.commit()

    total = len(records)
    print(f"  ✓ Inserted {total} MTY records into {table_name}")

    # Summary by month
    cursor = conn.execute(
        f"""
        SELECT month, COUNT(*) as line_items FROM {table_name}
        GROUP BY month
        ORDER BY CASE month
            WHEN 'Apr' THEN 1 WHEN 'May' THEN 2 WHEN 'Jun' THEN 3
            WHEN 'Jul' THEN 4 WHEN 'Aug' THEN 5 WHEN 'Sep' THEN 6
            WHEN 'Oct' THEN 7 WHEN 'Nov' THEN 8 WHEN 'Dec' THEN 9
            WHEN 'Jan' THEN 10 WHEN 'Feb' THEN 11 WHEN 'Mar' THEN 12
        END
    """
    )
    for month_name, line_items in cursor:
        print(f"    {month_name}: {line_items} line items")

    if wb_src is not None:
        wb_src.close()

    return total


if __name__ == "__main__":
    from sheets.db import get_connection

    conn = get_connection()
    count = calculate_mty(conn, "25_26", "/home/kbg/Documents/PriyaFil/12. MIS Mar 26.xlsx" )

    # Verify Dec output vs expected Excel values
    print("\nDec verification (compare to MISDec25Latest.xlsx MTY sheet):")
    print(f"{'Line Item':45s} {'Calculated':>15s}")
    print("-" * 62)
    cursor = conn.execute(
        """
        SELECT line_item, value FROM mty_25_26
        WHERE month = 'Dec'
        ORDER BY id
    """
    )
    rows: list[tuple[Any, Any]] = []
    if cursor is not None:
        rows = cast(list[tuple[Any, Any]], list(cursor))
    for line_item, value in rows:
        print(f"  {line_item:43s} {value:>15,.0f}")

    conn.close()

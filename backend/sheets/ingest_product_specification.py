"""
ingest_product_specification.py
================================
Ingest product specifications from 'product specification.xlsx' → Generation sheet.

Source: <workspace root>/product specification.xlsx → sheet 'Generation'

Column mapping (Excel → DB):
  Col 0: SI No                  → SKIP
  Col 1: Item master            → product_name
  Col 2: Code                   → product_code
  Col 3: Width                  → width
  Col 4: Length                 → length
  Col 5: Weight                 → weight
  Col 6: Mts/kg                 → mts_per_kg
  Col 7: GRM                    → grm
  Col 8: GSM                    → gsm
  Col 9: Fabrication charge     → fabrication_charge_basic

Calculated on ingestion:
  fabrication_charge_per_kg = fabrication_charge_basic * mts_per_kg

Table: product_specification  (no FY suffix — static reference data)
Re-running is safe: uses INSERT … ON CONFLICT DO UPDATE (upsert).
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional

import openpyxl

from .db import get_connection

SHEET_NAME = "Sheet1"
DEFAULT_FILE = "product specification.xlsx"


def create_product_specification_table(conn) -> None:
    conn.execute("""
        CREATE TABLE IF NOT EXISTS product_specification (
            id                        SERIAL  PRIMARY KEY,
            product_name              TEXT    NOT NULL,
            product_code              INTEGER,
            width                     REAL,
            length                    REAL,
            weight                    REAL,
            mts_per_kg                REAL,
            grm                       REAL,
            gsm                       REAL,
            fabrication_charge_basic  REAL,
            fabrication_charge_per_kg REAL,
            UNIQUE (product_name)
        )
    """)
    conn.commit()
    print("  ✓ Table ready: product_specification")


def safe_float(val) -> Optional[float]:
    if val is None or val == "" or val == " ":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_int(val) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def ingest_product_specification(
    conn,
    file_path: Path | None = None,
) -> int:
    """
    Ingest product specifications from the Generation sheet.

    Args:
        conn:      Database connection (psycopg2-compatible).
        file_path: Path to 'product specification.xlsx'.
                   Defaults to the workspace root (parent of sheets/).

    Returns:
        Number of rows upserted.
    """
    if file_path is None:
        file_path = Path(__file__).resolve().parents[1] / DEFAULT_FILE

    if not file_path.exists():
        raise FileNotFoundError(f"Product specification file not found: {file_path}")

    print(f"\n[PRODUCT SPECIFICATION] Reading from: {file_path.name}")
    create_product_specification_table(conn)

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{SHEET_NAME}' not found in {file_path.name}. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # row 1 = headers
    wb.close()

    records = []
    for row in rows:
        # Skip blank rows
        if not row or row[1] is None:
            continue

        product_name = str(row[1]).strip()
        if not product_name:
            continue

        product_code             = safe_int(row[2])
        width                    = safe_float(row[3])
        length                   = safe_float(row[4])
        weight                   = safe_float(row[5])
        mts_per_kg               = safe_float(row[6])
        grm                      = safe_float(row[7])
        gsm                      = safe_float(row[8])
        fabrication_charge_basic = safe_float(row[9])

        # Derived: fabrication cost per kg of product
        if fabrication_charge_basic is not None and mts_per_kg is not None:
            fabrication_charge_per_kg = round(fabrication_charge_basic * mts_per_kg, 4)
        else:
            fabrication_charge_per_kg = None

        records.append((
            product_name,
            product_code,
            width,
            length,
            weight,
            mts_per_kg,
            grm,
            gsm,
            fabrication_charge_basic,
            fabrication_charge_per_kg,
        ))

    conn.executemany(
        """
        INSERT INTO product_specification
            (product_name, product_code, width, length, weight,
             mts_per_kg, grm, gsm, fabrication_charge_basic, fabrication_charge_per_kg)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (product_name) DO UPDATE SET
            product_code              = EXCLUDED.product_code,
            width                     = EXCLUDED.width,
            length                    = EXCLUDED.length,
            weight                    = EXCLUDED.weight,
            mts_per_kg                = EXCLUDED.mts_per_kg,
            grm                       = EXCLUDED.grm,
            gsm                       = EXCLUDED.gsm,
            fabrication_charge_basic  = EXCLUDED.fabrication_charge_basic,
            fabrication_charge_per_kg = EXCLUDED.fabrication_charge_per_kg
        """,
        records,
    )
    conn.commit()
    print(f"  ✓ Upserted {len(records)} product specification records")
    return len(records)


if __name__ == "__main__":
    conn = get_connection()
    try:
        n = ingest_product_specification(conn)
        print(f"\nDone — {n} records in product_specification table")
    except Exception as exc:
        conn.rollback()
        print(f"❌ Failed: {exc}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()
